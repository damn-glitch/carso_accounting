import streamlit as st
import pandas as pd
from datetime import datetime, timedelta, timezone, date   # <-- добавили `timezone`
import calendar
import io
import psycopg2
from psycopg2 import sql
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import os
import logging
import requests
import threading
import time
from models.user import User  # Импортируем класс User из models/user.py
from models.service_record import ServiceRecord  # Импортируем класс ServiceRecord
from models.contract import Contract  # Импортируем класс Contract
from models.warranty_policy import WarrantyPolicy  # Импортируем класс WarrantyPolicy
from models.car import Car  # Импортируем класс Car

# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Настройка страницы
st.set_page_config(page_title="Учетная система автосалона", layout="wide")

# Система авторизации
VALID_USERS = {
    "manager1": "auto1111",
    "manager2": "auto1111",
    "manager3": "auto1111",
    "manager4": "auto1111",
    "leader": "alisher_krutoy"
}

KEYCLOAK_URL = "https://auth.carso.kz"


ACCESS_TOKEN = None  # Переменная для хранения токена доступа
# REFRESH_TOKEN = None  # Переменная для хранения токена обновления
ACCESS_TOKEN_EXPIRY = None  # Переменная для хранения времени истечения токена доступа
# REFRESH_TOKEN_EXPIRY = None  # Переменная для хранения времени истечения токена обновления

def check_login(username, password):
    """Проверка логина и пароля"""
    return username in VALID_USERS and VALID_USERS[username] == password

def is_leader(username):
    """Проверка, является ли пользователь руководителем"""
    return username == "leader"

def get_user_role(username):
    """Получение роли пользователя"""
    if is_leader(username):
        return "👑 Руководитель"
    else:
        return "👤 Менеджер"

def login_form():
    """Форма входа в систему"""
    st.markdown("""
    <div style="text-align: center; padding: 50px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius: 10px; margin-bottom: 30px;">
        <h1 style="color: white; margin: 0;">🚗 CARSO.KZ</h1>
        <h2 style="color: white; margin: 10px 0;">Учетная система автосалона</h2>
        <p style="color: #e0e0e0; margin: 0;">Система управления автомобильными продажами</p>
        <hr style="border-color: rgba(255,255,255,0.3); margin: 20px 0;">
        <p style="margin: 0; color: white; font-size: 16px;">
            💻 Сделал <strong>Алишер Бейсембеков</strong>, автор программы и учредитель Carso<br>
            <small style="opacity: 0.8;">© 2025 CARSO.KZ - Система управления автосалоном</small>
        </p>
    </div>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 2, 1])

    with col2:
        with st.container():
            st.markdown("### 🔐 Авторизация менеджера")
            username = st.text_input("👤 Логин", placeholder="Введите ваш логин (manager1-4)")
            password = st.text_input("🔒 Пароль", type="password", placeholder="Введите пароль")
            col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
            with col_btn2:
                if st.button("🚀 Войти в систему", type="primary", use_container_width=True):
                    if check_login(username, password):
                        st.session_state.authenticated = True
                        st.session_state.current_user = username
                        st.success(f"✅ Добро пожаловать, {username}!")
                        st.balloons()
                        st.rerun()
                    else:
                        st.error("❌ Неверный логин или пароль!")
                        st.warning("💡 Проверьте правильность ввода данных")
            st.markdown("---")
            st.markdown("""
            <div style="background-color: #f8f9fa; padding: 15px; border-radius: 8px; border-left: 4px solid #007bff;">
                <h4 style="margin-top: 0;">ℹ️ Информация для входа</h4>
                <p><strong>Менеджеры:</strong> manager1, manager2, manager3, manager4</p>
                <p><strong>Руководитель:</strong> leader</p>
                <p><strong>Пароли:</strong> у каждой роли свой пароль</p>
                <p><small>При проблемах со входом обратитесь к администратору системы</small></p>
            </div>
            """, unsafe_allow_html=True)

def logout():
    """Выход из системы"""
    st.session_state.authenticated = False
    st.session_state.current_user = None
    st.rerun()


def fetch_new_token():
    """Запрашиваем новый access_token по client_credentials."""
    global ACCESS_TOKEN, ACCESS_EXPIRES

    resp = requests.post(
        KEYCLOAK_URL + "/realms/carso/protocol/openid-connect/token",
        data={
            "grant_type":    "client_credentials",
            "client_id":     "carso_accounting",
            "client_secret": "qROypNhSAL9qoefJrxqU2k7PKUwJhxzK",
        }
    )
    resp.raise_for_status()
    data = resp.json()

    now = datetime.now(timezone.utc)
    ACCESS_TOKEN   = data["access_token"]
    expires_in     = data.get("expires_in", 3600)
    ACCESS_EXPIRES = now + timedelta(seconds=expires_in)
    print(f"[{now.isoformat()}] Новый токен действителен до {ACCESS_EXPIRES.isoformat()}")


def token_checker_loop():
    """Фон: проверяем раз в минуту, и при необходимости обновляем."""
    while True:
        now = datetime.now(timezone.utc)
        if ACCESS_TOKEN is None or (ACCESS_EXPIRES and now >= ACCESS_EXPIRES):
            print(f"[{now.isoformat()}] Токен истёк — запрашиваем новый")
            fetch_new_token()
        time.sleep(60)

# стартуем
fetch_new_token()
threading.Thread(target=token_checker_loop, daemon=True).start()

# def start_token_scheduler():
#     """Запускает фоновый тред единожды."""
#     if not threading.active_count() > 1:
#         threading.Thread(target=token_checker_loop, daemon=True).start()

# # Инициализация при старте
# fetch_new_tokens()
# start_token_scheduler()

def save_form_data():
    st.session_state.username = st.session_state.get("username", "")
    st.session_state.first_name = st.session_state.get("first_name", "")
    st.session_state.last_name = st.session_state.get("last_name", "")
    st.session_state.password = st.session_state.get("password", "")
    st.session_state.email = st.session_state.get("email", "")
    st.session_state.role = st.session_state.get("role", "Customer")
    st.session_state.full_name = st.session_state.get("full_name", "")
    st.session_state.iin = st.session_state.get("iin", "")

def add_user():
    st.header("Добавить нового пользователя")

    # Form fields with session state (to keep the data between reruns)
    username = st.text_input("Мобильный телефон пользователя", value=st.session_state.get("username", ""))
    first_name = st.text_input("Имя пользователя", value=st.session_state.get("first_name", ""))
    last_name = st.text_input("Фамилия пользователя", value=st.session_state.get("last_name", ""))
    password = st.text_input("Пароль", type="password", value=st.session_state.get("password", ""))
    email = st.text_input("Электронная почта", value=st.session_state.get("email", ""))
    role = st.selectbox("Роль пользователя", ["Customer", "Supplier"], index=["Customer", "Supplier"].index(st.session_state.get("role", "Customer")))
    full_name = st.text_input("Полное имя пользователя", value=st.session_state.get("full_name", ""))
    iin = st.text_input("ИИН пользователя", value=st.session_state.get("iin", ""))

    # Button to trigger user creation
    if st.button("Добавить пользователя"):
        # Collecting user data from the form
        user_data = {
            "username": username,
            "password": password,
            "email": email,
            "first_name": first_name,
            "last_name": last_name,
            "role": role,
            "full_name": full_name,
            "iin": iin,
        }

        # Call the create_user function
        created_user, error_message = create_user(user_data)
        
        # Show success or error message
        if created_user:
            st.success("Пользователь успешно добавлен!")
        else:
            st.error(f"Ошибка при добавлении пользователя: {error_message}")


# Function to create a new user in Keycloak
def create_user(data):
    global ACCESS_TOKEN
    url = KEYCLOAK_URL +"/admin/realms/carso/users"
    try:
        response = requests.post(
            url,
            headers={
                "Authorization": f"Bearer {ACCESS_TOKEN}",
                "Content-Type": "application/json"
            },
            json={
                "username": data["username"],  # username == mobile phone
                "email": data["email"],
                "firstName": data["first_name"],
                "lastName": data["last_name"],
                "enabled": True,
                "credentials": [{
                    "type": "password",
                    "value": data["password"],
                    "temporary": False
                }],
                "attributes": {
                    "iin": [data.get("iin", "")],
                    "full_name": [data.get("full_name", "")],
                    "mobile_number": ["34243242"]
                }
            }
        )

        # If request was successful, return True, else return the error message
        if response.status_code == 201:
            return True, None
        else:
            error_message = response.json()
            return False, error_message

    except requests.exceptions.RequestException as e:
        return False, {"error": "An error occurred", "message": str(e)}


def get_all_users():
    global ACCESS_TOKEN
    url = KEYCLOAK_URL + "/admin/realms/carso/users"  # Adjust URL based on your Keycloak setup

    resp = requests.get(url, headers={
        "Authorization": f"Bearer {ACCESS_TOKEN}",
        "Content-Type": "application/json"
    })
    resp.raise_for_status()
    raw = resp.json()  # List[dict]
    return [
        User(
            id=u.get("id"),
            username=u.get("username"),
            first_name=u.get("firstName"),
            last_name=u.get("lastName"),
            email=u.get("email"),
            enabled=u.get("enabled", False),
            attributes=u.get("attributes", {}),
        )
        for u in raw
    ]

def show_all_users():
    users = get_all_users()  # -> List[User]
    if not users:
        st.info("Пользователи не найдены.")
        return

    # Собираем мапу full_name -> User
    name_map = {u.full_name: u for u in users}
    full_names = list(name_map.keys())

    # 1) Инициализация стейта (делаем это один раз)
    if "selected_user" not in st.session_state:
        st.session_state.selected_user = full_names[0]
    if "show_details" not in st.session_state:
        st.session_state.show_details = False

    st.header("Список пользователей")

    # 2) Selectbox с ключом, совпадающим с именем в session_state
    selected_user = st.selectbox(
        "Выберите пользователя",
        full_names,
        index=full_names.index(st.session_state.selected_user),
        key="selected_user"    # <-- здесь key == имя переменной в session_state
    )

    # 3) Кнопки с on_click, чтобы править флаг, но не терять selectbox
    def show_flag():
        st.session_state.show_details = True

    def hide_flag():
        st.session_state.show_details = False

    st.button("Показать детали", on_click=show_flag)
    st.button("Скрыть детали", on_click=hide_flag)

    # 4) Отображаем подробности, если флаг поднят
    if st.session_state.show_details:
        u = name_map[selected_user]
        st.markdown(f"### {u.full_name}")
        st.write("• **ID**:",      u.id)
        st.write("• **Username**:", u.username)
        st.write("• **Email**:",    u.email)
        st.write("• **Enabled**:",  u.enabled)
        if u.attributes:
            st.write("• **Attributes**:")
            st.json(u.attributes)

        def delete_selected_user():
            url = KEYCLOAK_URL + "/admin/realms/carso/users/{u.id}"
            resp = requests.put(
                url,
                headers={
                    "Authorization": f"Bearer {ACCESS_TOKEN}",
                    "Content-Type": "application/json"
                },
                json={"enabled": False}
            )
            resp.raise_for_status()
            st.success(f"Пользователь «{u.full_name}» удалён (enabled=False).")
            # сброс флагов и выбор первого пользователя
            st.session_state.show_details = False
            st.session_state.selected_user = full_names[0]

        st.button(
            "Удалить пользователя",
            on_click=delete_selected_user
        )


def get_all_cars():
    global ACCESS_TOKEN
    url = "http://localhost:9080/cars" # Поменять когда запустим на проде

    resp = requests.get(url, headers={
        "Authorization": f"Bearer {ACCESS_TOKEN}",
        "Content-Type": "application/json"
    })
    resp.raise_for_status()
    raw = resp.json()  # List[dict]
    # Parse the response into a list of CarResponseDTO objects
    cars = [parse_car_response(car_data) for car_data in raw]

    return cars

# Streamlit code to create the button and display the cars
def show_all_cars():
    if st.button("Показать все автомобили"):
        cars = get_all_cars()  # Fetch the list of cars

        if not cars:
            st.info("Автомобили не найдены.")
            return

        for car in cars:
            st.subheader(f"Марка: {car.brand}, Модель: {car.model}, Год: {car.year}")
            st.write(f"• **VIN**: {car.vin}")
            st.write(f"• **ID**: {car.id}")
            
            if car.warranty_policy:
                st.write("• **Гарантия**:")
                st.write(f"  - ID: {car.warranty_policy.id}")
                st.write(f"  - Срок окончания: {car.warranty_policy.end_time}")
                st.write(f"  - Максимальный пробег: {car.warranty_policy.max_mileage} км")
            
            # Display service records if they exist
            if car.service_record_list:
                st.write("• **Сервисные записи**:")
                for record in car.service_record_list:
                    st.write(f"  - Тип: {record.service_type}, Пробег: {record.mileage} км")
                    st.write(f"    Описание: {record.description}")
            
            # Display contract details if they exist
            if car.contract_response_dto_list:
                st.write("• **Контракты**:")
                for contract in car.contract_response_dto_list:
                    st.write(f"  - Название контракта: {contract.contract_name}")
                    st.write(f"    Номер контракта: {contract.contract_number}")
                    st.write(f"    Статус: {contract.status}")
                    st.write(f"    Ссылка: {contract.link}")

def parse_car_response(car_data):
    # Assuming car_data is a dictionary containing a car's details
    warranty_policy = car_data.get("warrantyPolicy")
    warranty_policy_dto = None
    if warranty_policy:
        warranty_policy_dto = WarrantyPolicy(
            id=warranty_policy.get("id"),
            created_time=datetime.fromisoformat(warranty_policy.get("createdTime")),
            car_id=warranty_policy.get("carId"),
            end_time=datetime.fromisoformat(warranty_policy.get("endTime")),
            max_mileage=warranty_policy.get("maxMileage")
        )

    service_record_list = []
    if car_data.get("serviceRecordList"):
        for record in car_data["serviceRecordList"]:
            service_record_list.append(ServiceRecord(
                id=record.get("id"),
                mileage=record.get("mileage"),
                service_type=record.get("serviceType"),
                description=record.get("description"),
                car_id=record.get("carId"),
                service_center_id=record.get("serviceCenterId")
            ))

    contract_response_dto_list = []
    if car_data.get("contractResponseDTOList"):
        for contract in car_data["contractResponseDTOList"]:
            contract_response_dto_list.append(Contract(
                id=contract.get("id"),
                contract_name=contract.get("contractName"),
                contract_number=contract.get("contractNumber"),
                status=contract.get("status"),
                link=contract.get("link")
            ))

    # Now return the CarResponseDTO object
    return Car(
        id=car_data.get("id"),
        vin=car_data.get("vin"),
        brand=car_data.get("brand"),
        model=car_data.get("model"),
        year=car_data.get("year"),
        warranty_policy=warranty_policy_dto,
        service_record_list=service_record_list,
        contract_response_dto_list=contract_response_dto_list
    )

current_user = st.session_state.get('current_user', None)


# Проверка авторизации
if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    login_form()
    st.stop()


# # Now we can check if the user is a leader
# if current_user and is_leader(current_user):
#     with st.sidebar:
#         # Показать форму только один раз, если форма уже открыта
#         if "user_form_shown" not in st.session_state:
#             st.session_state.user_form_shown = False

#         # Проверяем, был ли клик по кнопке "Добавить нового пользователя"
#         if not st.session_state.user_form_shown:
#             if st.button("Добавить нового пользователя"):
#                 st.session_state.user_form_shown = True

#         # Показать форму для добавления пользователя, если была нажата кнопка
#         if st.session_state.user_form_shown:
#             add_user()
# Конфигурация машин и цен
CAR_TYPES = {
    "ГИБРИД 150К": 150000,
    "ГИБРИД 250К": 250000,
    "ГИБРИД 300К": 300000,
    "ДВС 70К": 70000,
    "БУУ АВТО 50К": 50000
}

# Цвета для типов машин
CAR_TYPE_COLORS = {
    "ГИБРИД 150К": "4472C4",
    "ГИБРИД 250К": "C5504B",
    "ГИБРИД 300К": "70AD47",
    "ДВС 70К": "FFC000",
    "БУУ АВТО 50К": "7030A0"
}

# Предустановленные автосалоны
DEFAULT_DEALERSHIPS = [
    "EL MOTORS", "LID CAR", "URBAN AUTO", "SAKURA MOTORS", "DREAM CAR",
    "NEW CAR/BOSS CAR", "CARMAX", "LYNG GO", "REAL AUTO ALMATY", "BOSSCAR",
    "GALAXY AVTO CENTER", "AST MOTORS", "ARB KAZAKHSTAN", "LYNK&CO AUTOINVEST",
    "ALMATY AUTO MOTORS", "HALYK AUTO", "СООРУЖЕНИЕ", "EV MASTER",
    "CHANGAN AUTO ALMATY", "DREAM CAR MARKET", "MIX AUTO", "MIR AUTO",
    "RAMADA AUTOMARKET", "SILK ROAD AUTO", "AVTOSHOPPING02", "CHINACARS",
    "ES MOTORS", "Q4TULPAR", "AVTOMARKET", "UCAR", "AUTO_CENTER_KZ",
    "ZENITH AUTO & DETAILING", "REGIONAUTOKZ"
]

# Инициализация базы данных PostgreSQL
@st.cache_resource(hash_funcs={psycopg2.extensions.connection: id})
def init_database():
    load_dotenv()
    DB_CONNECTION_STRING = os.getenv("DB_CONNECTION_STRING")
    if not DB_CONNECTION_STRING:
        st.error("Строка подключения к базе данных не найдена в .env файле!")
        raise ValueError("DB_CONNECTION_STRING is not set")

    try:
        conn = psycopg2.connect(DB_CONNECTION_STRING)
        conn.autocommit = True
        cursor = conn.cursor()

        # Создание таблицы автосалонов
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS dealerships (
                id SERIAL PRIMARY KEY,
                name TEXT UNIQUE NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Создание таблицы машин
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS cars (
                id SERIAL PRIMARY KEY,
                dealership_id INTEGER NOT NULL,
                car_type TEXT NOT NULL,
                count INTEGER NOT NULL,
                price_per_car INTEGER NOT NULL,
                total_amount INTEGER NOT NULL,
                date_added DATE NOT NULL,
                is_paid BOOLEAN DEFAULT FALSE,
                payment_date DATE,
                created_by TEXT,
                updated_by TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (dealership_id) REFERENCES dealerships(id) ON DELETE CASCADE
            )
        ''')

        # Добавление базовых автосалонов
        for dealership in DEFAULT_DEALERSHIPS:
            cursor.execute(
                sql.SQL('INSERT INTO dealerships (name) VALUES (%s) ON CONFLICT (name) DO NOTHING'),
                (dealership,)
            )

        cursor.execute('SELECT COUNT(*) FROM dealerships')
        existing_count = cursor.fetchone()[0]
        if existing_count < len(DEFAULT_DEALERSHIPS):
            st.info(f"Обновляем базу автосалонов... Добавлено {len(DEFAULT_DEALERSHIPS) - existing_count} новых автосалонов")

        cursor.close()
        logger.info("Successfully connected to PostgreSQL")
        return conn

    except psycopg2.Error as e:
        logger.error(f"Database connection error: {str(e)}")
        st.error(f"Ошибка подключения к PostgreSQL: {str(e)}")
        raise

# Функции для работы с БД
def get_dealerships(conn):
    cursor = conn.cursor()
    cursor.execute('SELECT id, name FROM dealerships ORDER BY name')
    result = cursor.fetchall()
    cursor.close()
    return result

def add_dealership(conn, name):
    cursor = conn.cursor()
    try:
        cursor.execute('INSERT INTO dealerships (name) VALUES (%s) ON CONFLICT (name) DO NOTHING', (name,))
        conn.commit()
        cursor.close()
        return True
    except psycopg2.Error:
        cursor.close()
        return False

def can_add_cars_for_dealership(conn, dealership_id, target_date):
    """Все автосалоны могут добавлять машины без предоплаты"""
    return True

def add_car_entry(conn, dealership_id, car_type, count, date_added, is_paid=False):
    cursor = conn.cursor()
    price_per_car = CAR_TYPES[car_type]
    total_amount = price_per_car * count
    current_user = st.session_state.get('current_user', 'unknown')
    payment_date = date.today() if is_paid else None
    updated_by = current_user if is_paid else None

    cursor.execute('''
        INSERT INTO cars (
            dealership_id, car_type, count, price_per_car, total_amount, date_added,
            is_paid, payment_date, created_by, updated_by
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    ''', (
        dealership_id, car_type, count, price_per_car, total_amount, date_added,
        is_paid, payment_date, current_user, updated_by
    ))
    conn.commit()
    cursor.close()

def update_car_payment_status(conn, car_id, is_paid):
    cursor = conn.cursor()
    current_user = st.session_state.get('current_user', 'unknown')
    cursor.execute('''
        UPDATE cars
        SET is_paid = %s,
            payment_date = %s,
            updated_by = %s
        WHERE id = %s
    ''', (is_paid, date.today() if is_paid else None, current_user, car_id))
    conn.commit()
    cursor.close()

def get_car_payment_status_for_today(conn, car_id):
    cursor = conn.cursor()
    cursor.execute('''
        SELECT is_paid, payment_date
        FROM cars
        WHERE id = %s
    ''', (car_id,))
    result = cursor.fetchone()
    cursor.close()
    if not result:
        return False
    is_paid, payment_date = result
    if not is_paid or not payment_date:
        return False
    payment_date = datetime.strptime(payment_date.strftime('%Y-%m-%d'), '%Y-%m-%d').date()
    return payment_date == date.today()

def get_cars_by_month_dealership(conn, year, month, dealership_id=None):
    cursor = conn.cursor()
    query = '''
        SELECT c.*, d.name as dealership_name
        FROM cars c
        JOIN dealerships d ON c.dealership_id = d.id
        WHERE EXTRACT(YEAR FROM c.date_added) = %s
          AND EXTRACT(MONTH FROM c.date_added) = %s
    '''
    params = [year, month]
    if dealership_id:
        query += ' AND c.dealership_id = %s'
        params.append(dealership_id)
    query += ' ORDER BY d.name, c.date_added'
    cursor.execute(query, params)
    result = cursor.fetchall()
    cursor.close()
    return result

def get_monthly_summary(conn, year, month):
    cursor = conn.cursor()
    cursor.execute('''
        SELECT 
            d.name AS dealership_name,
            c.car_type,
            SUM(c.count) AS total_count,
            SUM(c.total_amount) AS total_amount,
            SUM(CASE WHEN c.is_paid THEN c.count ELSE 0 END) AS paid_count,
            COUNT(CASE WHEN c.is_paid THEN 1 END) AS paid_entries
        FROM cars c
        JOIN dealerships d ON c.dealership_id = d.id
        WHERE EXTRACT(YEAR FROM c.date_added) = %s
          AND EXTRACT(MONTH FROM c.date_added) = %s
        GROUP BY d.id, d.name, c.car_type
        ORDER BY d.name, c.car_type
    ''', (year, month))
    result = cursor.fetchall()
    cursor.close()
    return result

def get_cars_by_day(conn, year, month, dealership_name, car_type):
    cursor = conn.cursor()
    cursor.execute('''
        SELECT EXTRACT(DAY FROM c.date_added) as day,
               SUM(c.count) as total_count
        FROM cars c
        JOIN dealerships d ON c.dealership_id = d.id
        WHERE EXTRACT(YEAR FROM c.date_added) = %s
          AND EXTRACT(MONTH FROM c.date_added) = %s
          AND d.name = %s
          AND c.car_type = %s
        GROUP BY EXTRACT(DAY FROM c.date_added)
    ''', (year, month, dealership_name, car_type))
    result = {}
    for row in cursor.fetchall():
        day = int(row[0])
        count = row[1]
        result[day] = count
    cursor.close()
    return result

import streamlit as st


# Функция создания Excel отчета
def create_excel_report(conn, year, month):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Отчет {calendar.month_name[month]} {year}"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    paid_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    unpaid_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    dealership_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    ws['C1'] = "ОТПРАВЛЕН СЧЕТ"
    ws['C1'].fill = unpaid_fill
    ws['C2'] = "СЧЕТ ОПЛАЧЕН"
    ws['C2'].fill = paid_fill

    ws['M3'] = "ТАБЛИЦА ОТЧЕТА CARSO.KZ"
    ws['M3'].font = header_font
    ws['M3'].fill = header_fill
    ws['M3'].alignment = center_alignment
    ws.merge_cells('M3:S3')

    ws['A4'] = "АВТОСАЛОН"
    ws['A4'].fill = dealership_fill
    ws['A4'].font = bold_font
    ws['A4'].alignment = center_alignment

    ws['B4'] = "ТИП МАШИНЫ"
    ws['B4'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    ws['B4'].font = Font(color="FFFFFF", bold=True)
    ws['B4'].alignment = center_alignment

    days_in_month = calendar.monthrange(year, month)[1]
    for day in range(1, days_in_month + 1):
        col = get_column_letter(3 + day - 1)
        ws[f'{col}4'] = day
        ws[f'{col}4'].fill = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")
        ws[f'{col}4'].alignment = center_alignment
        ws[f'{col}4'].font = bold_font

    total_col = get_column_letter(3 + days_in_month)
    paid_col = get_column_letter(3 + days_in_month + 1)

    ws[f'{total_col}4'] = "ВСЕГО"
    ws[f'{total_col}4'].fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
    ws[f'{total_col}4'].font = Font(color="FFFFFF", bold=True)
    ws[f'{total_col}4'].alignment = center_alignment

    ws[f'{paid_col}4'] = "ОПЛАЧЕНО"
    ws[f'{paid_col}4'].fill = paid_fill
    ws[f'{paid_col}4'].font = Font(color="FFFFFF", bold=True)
    ws[f'{paid_col}4'].alignment = center_alignment

    summary_data = get_monthly_summary(conn, year, month)
    current_row = 5
    dealership_groups = {}
    for row in summary_data:
        dealership = row[0]
        if dealership not in dealership_groups:
            dealership_groups[dealership] = []
        dealership_groups[dealership].append(row)

    for dealership, types in dealership_groups.items():
        start_row = current_row
        end_row = current_row + len(types) - 1
        if len(types) > 1:
            ws.merge_cells(f'A{start_row}:A{end_row}')
        ws[f'A{start_row}'] = dealership
        ws[f'A{start_row}'].fill = dealership_fill
        ws[f'A{start_row}'].font = Font(color="FFFFFF", bold=True)
        ws[f'A{start_row}'].alignment = center_alignment

        for car_data in types:
            _, car_type, total_count, total_amount, paid_count, _ = car_data
            ws[f'B{current_row}'] = car_type
            car_color = CAR_TYPE_COLORS.get(car_type, "808080")
            ws[f'B{current_row}'].fill = PatternFill(start_color=car_color, end_color=car_color, fill_type="solid")
            ws[f'B{current_row}'].font = Font(color="FFFFFF", bold=True)
            ws[f'B{current_row}'].alignment = center_alignment

            cars_by_day = get_cars_by_day(conn, year, month, dealership, car_type)
            for day, count in cars_by_day.items():
                col = get_column_letter(3 + day - 1)
                ws[f'{col}{current_row}'] = count
                ws[f'{col}{current_row}'].alignment = center_alignment

            ws[f'{total_col}{current_row}'] = total_count
            ws[f'{total_col}{current_row}'].alignment = center_alignment
            ws[f'{total_col}{current_row}'].font = bold_font

            ws[f'{paid_col}{current_row}'] = paid_count
            ws[f'{paid_col}{current_row}'].alignment = center_alignment
            if paid_count == total_count and total_count > 0:
                ws[f'{paid_col}{current_row}'].fill = paid_fill
            elif paid_count > 0:
                ws[f'{paid_col}{current_row}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

            current_row += 1
        current_row += 1

    legend_row = current_row + 2
    ws[f'A{legend_row}'] = "ЛЕГЕНДА ТИПОВ МАШИН:"
    ws[f'A{legend_row}'].font = bold_font
    legend_row += 1
    for idx, (car_type, price) in enumerate(CAR_TYPES.items()):
        col = get_column_letter(1 + idx)
        ws[f'{col}{legend_row}'] = f"{car_type} ({price:,} тг)"
        car_color = CAR_TYPE_COLORS.get(car_type, "808080")
        ws[f'{col}{legend_row}'].fill = PatternFill(start_color=car_color, end_color=car_color, fill_type="solid")
        ws[f'{col}{legend_row}'].font = Font(color="FFFFFF", bold=True)
        ws[f'{col}{legend_row}'].alignment = center_alignment

    return wb

# Инициализация БД
conn = init_database()


# Function to fetch admin access token


# Function to show the access token
def show_token():
    if st.button("Get Access Token"):
        access_token = get_admin_access_token()
        if access_token:
            st.success("Access Token Fetched Successfully!")
            st.markdown(f"### Your Access Token: {access_token}")
        else:
            st.error("Failed to fetch the access token.")



# Инициализация session state
if 'excel_reports' not in st.session_state:
    st.session_state.excel_reports = {}
if 'expanded_sections' not in st.session_state:
    st.session_state.expanded_sections = {}
if 'view_mode' not in st.session_state:
    current_user = st.session_state.get('current_user', 'unknown')
    st.session_state.view_mode = "По дням" if is_leader(current_user) else "По автосалонам"

# Заголовок с информацией о пользователе
header_col1, header_col2, header_col3 = st.columns([2, 2, 1])
with header_col1:
    st.title("🚗 Учетная система автосалона CARSO.KZ")
with header_col2:
    current_user = st.session_state.get('current_user', 'Неизвестно')
    user_role = get_user_role(current_user)
    st.markdown(f"""
    <div style="text-align: center; padding: 20px;">
        <h4>{user_role}: {current_user}</h4>
        <p>📅 {date.today().strftime('%d.%m.%Y')}</p>
    </div>
    """, unsafe_allow_html=True)
with header_col3:
    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("🚪 Выйти", type="secondary", help="Выход из системы"):
        logout()

st.divider()

if "show_all" not in st.session_state:
    st.session_state.show_all = False

if "show_all_cars" not in st.session_state:
    st.session_state.show_all_cars = False

# Show all users if the leader is logged in
if current_user and is_leader(current_user):
    with st.sidebar:
        st.header("👑 Панель руководителя")
        # Кнопки для управления флагом
        st.button(
            "Показать всех пользователей",
            on_click=lambda: st.session_state.__setitem__("show_all", True)
        )
        st.button(
            "Скрыть всех пользователей",
            on_click=lambda: st.session_state.__setitem__("show_all", False)
        )

          # Кнопки для управления состоянием отображения автомобилей
        st.button(
            "Показать все автомобили",
            on_click=lambda: st.session_state.update({"show_all_cars": True})
        )
        st.button(
            "Скрыть все автомобили",
            on_click=lambda: st.session_state.update({"show_all_cars": False})
        )


        # Показать форму только один раз, если форма уже открыта
        if "user_form_shown" not in st.session_state:
            st.session_state.user_form_shown = False

        # Проверяем, был ли клик по кнопке "Добавить нового пользователя"
        if not st.session_state.user_form_shown:
            if st.button("Добавить нового пользователя"):
                st.session_state.user_form_shown = True

        # Показать форму для добавления пользователя, если была нажата кнопка
        if st.session_state.user_form_shown:
            add_user()

if st.session_state.show_all:
    show_all_users()

# Если состояние show_all_cars = True, показываем все автомобили
if st.session_state.show_all_cars:
    show_all_cars()

# Боковая панель
with st.sidebar:
    current_user = st.session_state.get('current_user', 'Неизвестно')
    user_role = get_user_role(current_user)
    st.markdown(f"""
    <div style="background-color: #f0f2f6; padding: 10px; border-radius: 5px; margin-bottom: 20px;">
        <h4 style="margin: 0;">{user_role}</h4>
        <p style="margin: 0; font-size: 14px;"><strong>{current_user}</strong></p>
        <p style="margin: 0; font-size: 12px;">Активная сессия</p>
    </div>
    """, unsafe_allow_html=True)

    show_token()

    if is_leader(current_user):
        st.header("🏢 Управление автосалонами")
        total_dealerships = len(get_dealerships(conn))
        st.info(f"📊 Всего автосалонов в системе: **{total_dealerships}**")

        new_dealership = st.text_input("Новый автосалон")
        if st.button("Добавить автосалон"):
            if new_dealership:
                if add_dealership(conn, new_dealership):
                    st.success(f"✅ Добавлен автосалон: {new_dealership}")
                    st.rerun()
                else:
                    st.error("❌ Такой автосалон уже существует")

        st.divider()
        st.header("📊 Административная панель")
        cursor = conn.cursor()
        cursor.execute('SELECT SUM(count), SUM(total_amount) FROM cars')
        total_stats = cursor.fetchone()
        cursor.close()

        if total_stats[0]:
            col_a1, col_a2 = st.columns(2)
            with col_a1:
                st.metric("Всего машин", int(total_stats[0]))
            with col_a2:
                total_amount = int(total_stats[1])
                st.metric("Общий оборот", f"{total_amount:,} тг")

            cursor = conn.cursor()
            cursor.execute('SELECT SUM(count) FROM cars WHERE is_paid = TRUE')
            paid_cars = cursor.fetchone()[0] or 0
            cursor.close()

            if int(total_stats[0]) > 0:
                payment_rate = (paid_cars / int(total_stats[0])) * 100
                st.metric("Процент оплат", f"{payment_rate:.1f}%")

        st.divider()
        st.header("👥 Активность менеджеров")
        cursor = conn.cursor()
        cursor.execute('''
            SELECT created_by,
                   COUNT(*) AS entries_count,
                   SUM(count) AS total_cars,
                   SUM(total_amount) AS total_amount,
                   SUM(CASE WHEN is_paid THEN count ELSE 0 END) AS paid_cars
            FROM cars
            WHERE created_by IS NOT NULL AND created_by != 'unknown'
            GROUP BY created_by
            ORDER BY SUM(count) DESC
        ''')
        manager_detailed_stats = cursor.fetchall()
        cursor.close()

        if manager_detailed_stats:
            for manager, entries, cars_count, total_amount, paid_cars in manager_detailed_stats:
                cursor = conn.cursor()
                cursor.execute('SELECT SUM(count) FROM cars WHERE updated_by = %s AND is_paid = TRUE', (manager,))
                processed_payments = cursor.fetchone()[0] or 0
                cursor.close()
                efficiency = (paid_cars / cars_count * 100) if cars_count > 0 else 0
                with st.expander(f"📊 {manager} ({cars_count} машин)", expanded=False):
                    col_m1, col_m2, col_m3 = st.columns(3)
                    with col_m1:
                        st.metric("Записей", entries)
                        st.metric("Машин добавил", int(cars_count))
                    with col_m2:
                        st.metric("На сумму", f"{int(total_amount):,} тг")
                        st.metric("Оплачено машин", f"{paid_cars}/{cars_count}")
                    with col_m3:
                        st.metric("Обработал оплат", processed_payments)
                        st.metric("Эффективность", f"{efficiency:.1f}%")
        else:
            st.info("Статистика появится после работы менеджеров")
    else:
        st.header("Добавить машину")
        dealerships = get_dealerships(conn)
        dealership_dict = {name: id for id, name in dealerships}
        selected_dealership_name = st.selectbox("Автосалон", list(dealership_dict.keys()))
        selected_dealership_id = dealership_dict.get(selected_dealership_name)
        selected_date = st.date_input("Дата добавления", value=date.today())
        car_type = st.selectbox("Тип машины", list(CAR_TYPES.keys()))
        car_count = st.number_input("Количество машин", min_value=1, value=1)
        is_paid = st.checkbox("Машины оплачены", value=False, help="Отметьте если машины уже оплачены клиентом")

        if st.button("Добавить машины"):
            add_car_entry(conn, selected_dealership_id, car_type, car_count, selected_date, is_paid)
            status = "оплаченных" if is_paid else "неоплаченных"
            st.success(f"✅ Добавлено {car_count} {status} машин типа {car_type} для {selected_dealership_name}")
            st.rerun()

        st.divider()
        st.subheader("📈 Ваша статистика")
        cursor = conn.cursor()
        cursor.execute('SELECT SUM(count), SUM(total_amount) FROM cars WHERE created_by = %s', (current_user,))
        user_stats = cursor.fetchone()
        cursor.close()

        if user_stats[0]:
            st.metric("Добавлено машин", int(user_stats[0]))
            st.metric("На сумму", f"{int(user_stats[1]):,} тг")
        else:
            st.info("Вы еще не добавляли машины")

        cursor = conn.cursor()
        cursor.execute('SELECT SUM(count) FROM cars WHERE updated_by = %s AND is_paid = TRUE', (current_user,))
        user_payments = cursor.fetchone()[0] or 0
        cursor.close()
        if user_payments > 0:
            st.metric("Обработано оплат", f"{user_payments} машин")

    st.divider()
    if st.button("🚪 Выйти из системы", type="secondary", use_container_width=True):
        logout()

# Основной контент
col1, col2 = st.columns([3, 1])
with col1:
    st.header("📊 Учет по месяцам и автосалонам")
    view_col1, view_col2, view_col3, view_col4 = st.columns([2, 2, 2, 1])
    with view_col1:
        view_year = st.selectbox("Год для просмотра", range(2023, 2030), index=2)
    with view_col2:
        view_month = st.selectbox("Месяц для просмотра", range(1, 13), index=date.today().month - 1)
    with view_col3:
        view_mode_options = ["По автосалонам", "По дням"]
        current_index = view_mode_options.index(st.session_state.view_mode)
        selected_mode = st.selectbox("Группировка", view_mode_options, index=current_index)
        if selected_mode != st.session_state.view_mode:
            st.session_state.view_mode = selected_mode
            st.rerun()
        view_mode = st.session_state.view_mode
    with view_col4:
        if st.button("🔄 Обновить все", help="Обновить статус всех оплат"):
            st.rerun()

    month_name = calendar.month_name[view_month]
    cars_data = get_cars_by_month_dealership(conn, view_year, view_month)

    if cars_data:
        st.subheader(f"📅 {month_name} {view_year}")
        if view_mode == "По дням":
            day_groups = {}
            for car in cars_data:
                car_date = car[6]
                if car_date not in day_groups:
                    day_groups[car_date] = []
                day_groups[car_date].append(car)
            sorted_days = sorted(day_groups.keys(), reverse=True)

            for car_date in sorted_days:
                cars = day_groups[car_date]
                date_obj = datetime.strptime(str(car_date), '%Y-%m-%d').date()
                day_name = date_obj.strftime('%d %B %Y (%A)')
                day_total_cars = sum(car[3] for car in cars)
                day_total_amount = sum(car[5] for car in cars)
                day_paid_cars = sum(car[3] for car in cars if get_car_payment_status_for_today(conn, car[0]) or car[7])

                if date_obj == date.today():
                    bg_color = "#e8f5e8"
                    day_emoji = "📅"
                elif date_obj == date.today() - timedelta(days=1):
                    bg_color = "#fff3e0"
                    day_emoji = "📋"
                else:
                    bg_color = "#f8f9fa"
                    day_emoji = "📄"

                with st.container():
                    st.markdown(f"""
                    <div style="background-color: {bg_color}; padding: 15px; border-radius: 10px; margin: 10px 0;">
                        <h4>{day_emoji} {day_name}</h4>
                        <p>Машин: {day_total_cars} | Сумма: {day_total_amount:,} тг | Оплачено: {day_paid_cars}/{day_total_cars}</p>
                    </div>
                    """, unsafe_allow_html=True)

                    day_dealership_groups = {}
                    for car in cars:
                        dealership = car[12]
                        if dealership not in day_dealership_groups:
                            day_dealership_groups[dealership] = []
                        day_dealership_groups[dealership].append(car)

                    for dealership, dealership_cars in day_dealership_groups.items():
                        with st.expander(f"🏢 {dealership} ({len(dealership_cars)} записей)", expanded=False):
                            for car in dealership_cars:
                                car_id, dealership_id, car_type, count, price_per_car, total_amount, date_added, is_paid, payment_date, created_by, updated_by, created_at, dealership_name = car
                                paid_today = get_car_payment_status_for_today(conn, car_id)
                                if paid_today:
                                    status_color = "🟢"
                                    status_text = f"Оплачено сегодня ({updated_by or 'N/A'})" if updated_by else "Оплачено сегодня"
                                elif is_paid:
                                    status_color = "🟡"
                                    payment_info = f" ({updated_by})" if updated_by else ""
                                    status_text = f"Оплачено {payment_date}{payment_info}" if payment_date else f"Оплачено ранее{payment_info}"
                                else:
                                    status_color = "🔴"
                                    status_text = "Не оплачено"

                                car_col1, car_col2, car_col3, car_col4 = st.columns([2, 1, 1, 1])
                                with car_col1:
                                    creator_info = f" (добавил: {created_by})" if created_by else ""
                                    st.write(f"**{car_type}** - {count} шт.{creator_info}")
                                with car_col2:
                                    st.write(f"{total_amount:,} тг")
                                with car_col3:
                                    st.write(f"{status_color} {status_text}")
                                with car_col4:
                                    button_key = f"payment_btn_{car_id}_{date.today()}_{car_date}"
                                    if paid_today:
                                        if st.button("❌ Отменить оплату", key=f"unpay_{button_key}", type="secondary"):
                                            update_car_payment_status(conn, car_id, False)
                                            st.success("✅ Оплата отменена!")
                                            st.rerun()
                                    else:
                                        if st.button("✅ Оплатить сегодня", key=f"pay_{button_key}", type="primary"):
                                            update_car_payment_status(conn, car_id, True)
                                            st.success("✅ Оплата зарегистрирована на сегодня!")
                                            st.rerun()
        else:
            dealership_groups = {}
            for car in cars_data:
                dealership = car[12]
                if dealership not in dealership_groups:
                    dealership_groups[dealership] = []
                dealership_groups[dealership].append(car)

            for dealership, cars in dealership_groups.items():
                dealership_col1, dealership_col2 = st.columns([3, 1])
                with dealership_col2:
                    if st.button("🔄 Обновить", key=f"refresh_{dealership}_{view_year}_{view_month}", help="Обновить статус оплат"):
                        st.rerun()
                with st.expander(f"🏢 {dealership}", expanded=True):
                    for car in cars:
                        car_id, dealership_id, car_type, count, price_per_car, total_amount, date_added, is_paid, payment_date, created_by, updated_by, created_at, dealership_name = car
                        paid_today = get_car_payment_status_for_today(conn, car_id)
                        if paid_today:
                            status_color = "🟢"
                            status_text = f"Оплачено сегодня ({updated_by or 'N/A'})" if updated_by else "Оплачено сегодня"
                        elif is_paid:
                            status_color = "🟡"
                            payment_info = f" ({updated_by})" if updated_by else ""
                            status_text = f"Оплачено {payment_date}{payment_info}" if payment_date else f"Оплачено ранее{payment_info}"
                        else:
                            status_color = "🔴"
                            status_text = "Не оплачено"

                        car_col1, car_col2, car_col3, car_col4 = st.columns([2, 1, 1, 1])
                        with car_col1:
                            creator_info = f" (добавил: {created_by})" if created_by else ""
                            st.write(f"**{car_type}** - {count} шт. ({date_added}){creator_info}")
                        with car_col2:
                            st.write(f"{total_amount:,} тг")
                        with car_col3:
                            st.write(f"{status_color} {status_text}")
                        with car_col4:
                            button_key = f"payment_btn_{car_id}_{date.today()}_{view_year}_{view_month}"
                            if paid_today:
                                if st.button("❌ Отменить оплату", key=f"unpay_{button_key}", type="secondary"):
                                    update_car_payment_status(conn, car_id, False)
                                    st.success("✅ Оплата отменена!")
                                    st.rerun()
                            else:
                                if st.button("✅ Оплатить сегодня", key=f"pay_{button_key}", type="primary"):
                                    update_car_payment_status(conn, car_id, True)
                                    st.success("✅ Оплата зарегистрирована на сегодня!")
                                    st.rerun()

                    dealership_total = sum(car[5] for car in cars)
                    dealership_cars = sum(car[3] for car in cars)
                    paid_cars_today = sum(car[3] for car in cars if get_car_payment_status_for_today(conn, car[0]))
                    st.markdown("---")
                    summary_col1, summary_col2, summary_col3 = st.columns(3)
                    with summary_col1:
                        st.metric("Всего машин", dealership_cars)
                    with summary_col2:
                        st.metric("Оплачено сегодня", f"{paid_cars_today}/{dealership_cars}")
                    with summary_col3:
                        st.metric("Сумма", f"{dealership_total:,} тг")

        st.divider()
        export_col1, export_col2 = st.columns(2)
        with export_col1:
            if st.button(f"📊 Создать отчет Excel", key=f"create_excel_{view_year}_{view_month}"):
                wb = create_excel_report(conn, view_year, view_month)
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                st.session_state[f'excel_data_{view_year}_{view_month}'] = excel_buffer.getvalue()
                st.success("✅ Отчет создан! Нажмите кнопку скачивания справа.")
        with export_col2:
            excel_key = f'excel_data_{view_year}_{view_month}'
            if excel_key in st.session_state:
                st.download_button(
                    label=f"💾 Скачать отчет {month_name} {view_year}.xlsx",
                    data=st.session_state[excel_key],
                    file_name=f"Отчет_{month_name}_{view_year}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"download_{view_year}_{view_month}"
                )
            else:
                st.button("💾 Сначала создайте отчет", disabled=True)
    else:
        st.info(f"Данные за {month_name} {view_year} отсутствуют.")

with col2:
    if is_leader(current_user):
        st.header("👑 Панель руководителя")
        cursor = conn.cursor()
        cursor.execute('SELECT SUM(count), SUM(total_amount) FROM cars')
        total_stats = cursor.fetchone()
        cursor.close()

        if total_stats[0]:
            st.subheader("🎯 Ключевые показатели")
            total_cars = int(total_stats[0])
            total_revenue = int(total_stats[1])

            cursor = conn.cursor()
            cursor.execute('SELECT SUM(count) FROM cars WHERE is_paid = TRUE')
            paid_cars = cursor.fetchone()[0] or 0
            cursor.execute('SELECT SUM(total_amount) FROM cars WHERE is_paid = TRUE')
            paid_revenue = cursor.fetchone()[0] or 0
            cursor.close()

            col_kpi1, col_kpi2 = st.columns(2)
            with col_kpi1:
                st.metric("💰 Получено", f"{paid_revenue:,} тг")
                st.metric("🚗 Оплачено машин", f"{paid_cars}/{total_cars}")
            with col_kpi2:
                pending_revenue = total_revenue - paid_revenue
                st.metric("⏳ Ожидается", f"{pending_revenue:,} тг")
                payment_rate = (paid_cars / total_cars * 100) if total_cars > 0 else 0
                st.metric("📊 Конверсия", f"{payment_rate:.1f}%")
            st.progress(payment_rate / 100, text=f"Выполнение плана: {payment_rate:.1f}%")

            st.divider()
            st.subheader("📈 Динамика продаж")
            cursor = conn.cursor()
            cursor.execute('''
                SELECT date_added,
                       SUM(count) AS daily_cars,
                       SUM(total_amount) AS daily_revenue,
                       SUM(CASE WHEN is_paid THEN total_amount ELSE 0 END) AS daily_paid
                FROM cars
                WHERE date_added >= CURRENT_DATE - INTERVAL '30 days'
                GROUP BY date_added
                ORDER BY date_added
            ''')
            daily_data = cursor.fetchall()
            cursor.close()

            if daily_data:
                dates = [row[0] for row in daily_data]
                cars_data = [row[1] for row in daily_data]
                revenue_data = [row[2] for row in daily_data]
                paid_data = [row[3] for row in daily_data]
                chart_data = pd.DataFrame({
                    'Дата': dates,
                    'Добавлено машин': cars_data,
                    'Оборот (тыс. тг)': [x / 1000 for x in revenue_data],
                    'Получено (тыс. тг)': [x / 1000 for x in paid_data]
                })
                st.line_chart(chart_data.set_index('Дата'))

            st.divider()
            st.subheader("🚗 Популярность типов машин")
            cursor = conn.cursor()
            cursor.execute('''
                SELECT car_type,
                       SUM(count) AS total_count,
                       SUM(total_amount) AS total_amount,
                       SUM(CASE WHEN is_paid THEN count ELSE 0 END) AS paid_count
                FROM cars
                GROUP BY car_type
                ORDER BY SUM(count) DESC
            ''')
            car_types_data = cursor.fetchall()
            cursor.close()

            if car_types_data:
                types_chart = pd.DataFrame({
                    'Тип машины': [row[0] for row in car_types_data],
                    'Продано': [row[1] for row in car_types_data],
                    'Оплачено': [row[3] for row in car_types_data]
                })
                st.bar_chart(types_chart.set_index('Тип машины'))
                for car_type, total_count, total_amount, paid_count in car_types_data:
                    conversion = (paid_count / total_count * 100) if total_count > 0 else 0
                    with st.expander(f"🚙 {car_type}: {total_count} шт.", expanded=False):
                        col_ct1, col_ct2, col_ct3 = st.columns(3)
                        with col_ct1:
                            st.metric("Всего", total_count)
                        with col_ct2:
                            st.metric("Оплачено", f"{paid_count}/{total_count}")
                        with col_ct3:
                            st.metric("Конверсия", f"{conversion:.1f}%")

            st.divider()
            st.subheader("🏆 Рейтинг эффективности")
            cursor = conn.cursor()
            cursor.execute('''
                SELECT created_by,
                       COUNT(*) AS entries_count,
                       SUM(count) AS total_cars,
                       SUM(total_amount) AS total_amount,
                       SUM(CASE WHEN is_paid THEN count ELSE 0 END) AS paid_cars,
                       SUM(CASE WHEN is_paid THEN total_amount ELSE 0 END) AS paid_amount
                FROM cars
                WHERE created_by IS NOT NULL AND created_by != 'unknown'
                GROUP BY created_by
                ORDER BY SUM(CASE WHEN is_paid THEN total_amount ELSE 0 END) DESC
            ''')
            managers_rating = cursor.fetchall()
            cursor.close()

            if managers_rating:
                for i, (manager, entries, cars_count, total_amount, paid_cars, paid_amount) in enumerate(managers_rating, 1):
                    cursor = conn.cursor()
                    cursor.execute('SELECT SUM(count) FROM cars WHERE updated_by = %s AND is_paid = TRUE', (manager,))
                    processed_payments = cursor.fetchone()[0] or 0
                    cursor.close()
                    efficiency = (paid_cars / cars_count * 100) if cars_count > 0 else 0
                    if i == 1:
                        medal = "🥇"
                    elif i == 2:
                        medal = "🥈"
                    elif i == 3:
                        medal = "🥉"
                    else:
                        medal = f"{i}."
                    with st.expander(f"{medal} {manager} - {paid_amount:,} тг получено", expanded=False):
                        col_mr1, col_mr2, col_mr3 = st.columns(3)
                        with col_mr1:
                            st.metric("Добавил", f"{cars_count} машин")
                            st.metric("Записей", entries)
                        with col_mr2:
                            st.metric("Обработал оплат", processed_payments)
                            st.metric("Эффективность", f"{efficiency:.1f}%")
                        with col_mr3:
                            st.metric("Получил оплат", f"{paid_amount:,} тг")
                            avg_per_car = paid_amount / paid_cars if paid_cars > 0 else 0
                            st.metric("Средний чек", f"{avg_per_car:,.0f} тг")

            st.divider()
            st.subheader("🎯 Цели и прогнозы")
            current_month = date.today().month
            current_year = date.today().year
            cursor = conn.cursor()
            cursor.execute('''
                SELECT SUM(count) AS month_cars,
                       SUM(total_amount) AS month_revenue,
                       SUM(CASE WHEN is_paid THEN total_amount ELSE 0 END) AS month_paid
                FROM cars
                WHERE EXTRACT(YEAR FROM date_added) = %s
                  AND EXTRACT(MONTH FROM date_added) = %s
            ''', (current_year, current_month))
            month_stats = cursor.fetchone()
            cursor.close()

            if month_stats and month_stats[0]:
                month_cars, month_revenue, month_paid = month_stats
                days_passed = date.today().day
                days_in_current_month = calendar.monthrange(current_year, current_month)[1]
                days_remaining = days_in_current_month - days_passed
                if days_passed > 0:
                    daily_avg_revenue = month_revenue / days_passed
                    projected_revenue = daily_avg_revenue * days_in_current_month
                    col_pr1, col_pr2 = st.columns(2)
                    with col_pr1:
                        st.metric("📅 Дней осталось", days_remaining)
                        st.metric("📊 Средний оборот/день", f"{daily_avg_revenue:,.0f} тг")
                    with col_pr2:
                        st.metric("🎯 Прогноз на месяц", f"{projected_revenue:,.0f} тг")
                        month_progress = (days_passed / days_in_current_month) * 100
                        st.metric("⏰ Прогресс месяца", f"{month_progress:.1f}%")
                    revenue_progress = (month_revenue / projected_revenue * 100) if projected_revenue > 0 else 0
                    st.progress(month_progress / 100, text=f"Выполнение месячного плана: {revenue_progress:.1f}%")

            st.divider()
            st.subheader("🔥 Лучшие дни")
            cursor = conn.cursor()
            cursor.execute('''
                SELECT date_added,
                       SUM(count) AS daily_cars,
                       SUM(total_amount) AS daily_revenue
                FROM cars
                WHERE date_added >= CURRENT_DATE - INTERVAL '60 days'
                GROUP BY date_added
                ORDER BY SUM(total_amount) DESC
                LIMIT 5
            ''')
            top_days = cursor.fetchall()
            cursor.close()

            if top_days:
                for i, (day_date, cars, revenue) in enumerate(top_days, 1):
                    day_name = datetime.strptime(day_date.strftime('%Y-%m-%d'), '%Y-%m-%d').strftime('%d.%m.%Y (%A)')
                    col_td1, col_td2 = st.columns([3, 1])
                    with col_td1:
                        st.write(f"**{i}. {day_name}**")
                    with col_td2:
                        st.write(f"**{revenue:,} тг** ({cars} машин)")
    else:
        st.header("📊 Общая статистика")
        cursor = conn.cursor()
        cursor.execute('SELECT SUM(count), SUM(total_amount) FROM cars')
        total_stats = cursor.fetchone()
        cursor.close()

        if total_stats[0]:
            st.metric("Всего машин в системе", int(total_stats[0]))
            st.metric("Общая сумма", f"{int(total_stats[1]):,} тг")
            cursor = conn.cursor()
            cursor.execute('SELECT SUM(count) FROM cars WHERE is_paid = TRUE')
            paid_cars = cursor.fetchone()[0] or 0
            cursor.execute('SELECT SUM(total_amount) FROM cars WHERE is_paid = TRUE')
            paid_amount = cursor.fetchone()[0] or 0
            cursor.close()
            unpaid_cars = int(total_stats[0]) - paid_cars

            st.divider()
            st.subheader("💰 Статус оплат")
            st.caption(f"📅 За сегодня ({date.today().strftime('%d.%m.%Y')})")
            today_str = date.today().strftime('%Y-%m-%d')
            cursor = conn.cursor()
            cursor.execute('''
                SELECT SUM(count), SUM(total_amount)
                FROM cars
                WHERE payment_date = %s AND is_paid = TRUE
            ''', (today_str,))
            today_paid_stats = cursor.fetchone()
            today_paid_cars = today_paid_stats[0] or 0
            today_paid_amount = today_paid_stats[1] or 0
            cursor.execute('''
                SELECT SUM(count), SUM(total_amount)
                FROM cars
                WHERE date_added = %s AND is_paid = FALSE
            ''', (today_str,))
            today_unpaid_stats = cursor.fetchone()
            today_unpaid_cars = today_unpaid_stats[0] or 0
            today_unpaid_amount = today_unpaid_stats[1] or 0
            cursor.close()

            today_total_cars = today_paid_cars + today_unpaid_cars
            today_total_amount = today_paid_amount + today_unpaid_amount

            if today_total_cars > 0:
                col_today1, col_today2 = st.columns(2)
                with col_today1:
                    st.markdown("""
                    <div style="background-color: #d4edda; padding: 10px; border-radius: 8px; margin-bottom: 10px;">
                        <div style="display: flex; align-items: center;">
                            <span style="color: #28a745; font-size: 20px; margin-right: 8px;">✅</span>
                            <span style="color: #155724; font-weight: bold;">Оплачено</span>
                        </div>
                        <div style="font-size: 24px; font-weight: bold; color: #155724; margin: 5px 0;">
                            {today_paid_cars} машин
                        </div>
                        <div style="color: #6c757d; font-size: 14px;">
                            Сумма оплат<br>
                            <span style="font-size: 18px; font-weight: bold; color: #155724;">
                                {today_paid_amount:,} тг
                            </span>
                        </div>
                    </div>
                    """.format(today_paid_cars=today_paid_cars, today_paid_amount=today_paid_amount), unsafe_allow_html=True)
                with col_today2:
                    st.markdown("""
                    <div style="background-color: #f8d7da; padding: 10px; border-radius: 8px; margin-bottom: 10px;">
                        <div style="display: flex; align-items: center;">
                            <span style="color: #dc3545; font-size: 20px; margin-right: 8px;">❌</span>
                            <span style="color: #721c24; font-weight: bold;">Не оплачено</span>
                        </div>
                        <div style="font-size: 24px; font-weight: bold; color: #721c24; margin: 5px 0;">
                            {today_unpaid_cars} машин
                        </div>
                        <div style="color: #6c757d; font-size: 14px;">
                            К доплате<br>
                            <span style="font-size: 18px; font-weight: bold; color: #721c24;">
                                {today_unpaid_amount:,} тг
                            </span>
                        </div>
                    </div>
                    """.format(today_unpaid_cars=today_unpaid_cars, today_unpaid_amount=today_unpaid_amount), unsafe_allow_html=True)
                today_payment_percentage = (today_paid_cars / today_total_cars * 100) if today_total_cars > 0 else 0
                st.progress(today_payment_percentage / 100, text=f"Оплачено: {today_payment_percentage:.1f}%")
            else:
                st.info("📋 За сегодня машин не добавлялось")

            st.divider()
            st.subheader("🏢 Популярные автосалоны")
            cursor = conn.cursor()
            cursor.execute('''
                SELECT d.name, SUM(c.count), SUM(c.total_amount)
                FROM cars c
                JOIN dealerships d ON c.dealership_id = d.id
                GROUP BY d.id, d.name
                ORDER BY SUM(c.count) DESC LIMIT 5
            ''')
            top_dealerships = cursor.fetchall()
            cursor.close()

            for i, (name, total_cars, total_amount) in enumerate(top_dealerships, 1):
                st.write(f"**{i}. {name}** - {total_cars} машин ({total_amount:,} тг)")

            st.divider()
            st.subheader("🚗 Популярные типы машин")
            cursor = conn.cursor()
            cursor.execute('''
                SELECT car_type, SUM(count) as total_count
                FROM cars
                GROUP BY car_type
                ORDER BY SUM(count) DESC LIMIT 5
            ''')
            popular_types = cursor.fetchall()
            cursor.close()

            for i, (car_type, count) in enumerate(popular_types, 1):
                st.write(f"**{i}. {car_type}** - {count} машин")
        else:
            st.info("📈 Статистика появится после добавления машин")

st.markdown("---")
st.markdown("""
<div style="
    text-align: center; 
    padding: 25px; 
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
    border-radius: 15px; 
    margin: 30px 0;
    box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
">
    <p style="margin: 0; color: white; font-size: 16px;">
        💻 Сделал <strong>Алишер Бейсембеков</strong>, ген. директор и учредитель Carso<br>
        🎯 По концепции <strong>Санжар Тургали</strong>, региональный директор Carso<br>
        <small style="opacity: 0.8;">© 2025 CARSO.KZ - Система управления автосалоном</small>
    </p>
</div>
""", unsafe_allow_html=True)

if is_leader(current_user):
    st.markdown("---")
    st.subheader("⚠️ Административные функции")
    if st.button("🗑️ Очистить все данные", type="secondary", help="Очистка с восстановлением автосалонов"):
        if st.button("Подтвердить очистку всех данных", type="primary"):
            cursor = conn.cursor()
            cursor.execute('DELETE FROM cars')
            cursor.execute('DELETE FROM dealerships')
            for dealership in DEFAULT_DEALERSHIPS:
                cursor.execute('INSERT INTO dealerships (name) VALUES (%s) ON CONFLICT (name) DO NOTHING', (dealership,))
            conn.commit()
            cursor.close()
            st.success("Все данные очищены!")
            st.rerun()

    st.markdown("---")
    st.markdown("""
    <div style="background-color: #ffebee; border: 2px solid #f44336; border-radius: 8px; padding: 15px; margin: 10px 0;">
        <h4 style="color: #d32f2f; margin-top: 0;">🚨 ОПАСНАЯ ЗОНА</h4>
        <p style="color: #d32f2f; margin: 0;">Полная очистка базы данных без восстановления</p>
    </div>
    """, unsafe_allow_html=True)
    destroy_password = st.text_input(
        "🔐 Пароль для полной очистки",
        type="password",
        placeholder="Введите пароль для подтверждения",
        help="Требуется специальный пароль для полной очистки"
    )
    if st.button("💥 ПОЛНАЯ ОЧИСТКА БАЗЫ ДАННЫХ", type="primary", help="ВНИМАНИЕ: Полная очистка без восстановления!"):
        if destroy_password == "alisher_destroy":
            if st.button("🔥 ПОДТВЕРДИТЬ ПОЛНОЕ УНИЧТОЖЕНИЕ", type="primary"):
                cursor = conn.cursor()
                cursor.execute('DELETE FROM cars')
                cursor.execute('DELETE FROM dealerships')
                conn.commit()
                cursor.close()
                st.success("💀 База данных полностью очищена!")
                st.warning("⚠️ Все автосалоны удалены! Потребуется ручное восстановление.")
                st.rerun()
        elif destroy_password:
            st.error("❌ Неверный пароль для полной очистки!")
        else:
            st.warning("⚠️ Введите пароль для подтверждения полной очистки")
